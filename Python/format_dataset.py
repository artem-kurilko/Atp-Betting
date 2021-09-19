"""
Building of the raw dataset

Importation of the Excel files - 1 per year
Some preprocessing is necessary because for several years the ranking are not present.
Store data in this format: id, date, wName, lName, wRank, lRank, wElo, lElo, wAmountOfAllWins, lAmountOfAllWins
"""

import os
from pathlib import Path

import openpyxl


def build_dataset():
    # Get all raw data files
    raw_data_path = 'E:/practice/AtpPredictionBot'
    os.chdir(raw_data_path + '/Data')
    filenames = os.listdir('.')

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_dataset_index = 1

    # Read raw dataset and transfer player's names, date and ranking to train dataset
    for file in filenames:
        excel_file = Path(raw_data_path + '/Data', file)
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        rows = sheet.max_row

        # Retrieve cells values from dataset and write them into output file
        for row in range(2, rows + 1):
            try:
                date = sheet['D' + str(row)].value
                winner = sheet['J' + str(row)].value
                loser = sheet['K' + str(row)].value
                winner_rank = sheet['L' + str(row)].value
                loser_rank = sheet['M' + str(row)].value
                winner_coef = sheet['AC' + str(row)].value
                loser_coef = sheet['AD' + str(row)].value
            except:
                continue

            data = [output_dataset_index, date, winner, loser, winner_coef, loser_coef, winner_rank, loser_rank]
            output_sheet.append(data)
            output_dataset_index += 1

    # Save generated data to atp_data file
    os.chdir('E:/practice/AtpPredictionBot/Generated Data')
    output_workbook.save('atp_data.xlsx')


def add_total_wins_and_elo_to_dataset():
    """
    Add total amount of wins and elo to train dataset
    """
    # Prepare atp_data file
    raw_data_path = 'E:/practice/AtpPredictionBot/Generated Data'
    excel_file = Path(raw_data_path, 'atp_data.xlsx')
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    rows = sheet.max_row

    # Prepare features data file
    features_file = open('E:/practice/AtpPredictionBot/Generated Data/playersAmountOfWinsAndElo.txt', 'r')
    file_lines = features_file.readlines()

    for row in range(1, rows + 1):
        for line in file_lines:
            winner = sheet['C' + str(row)].value
            loser = sheet['D' + str(row)].value

            line_values = line.split(',')
            player = line_values[0].strip()
            amount_of_wins = line_values[1].strip()
            elo_ranking = line_values[2].strip()

            if winner == player:
                sheet['I' + str(row)] = int(amount_of_wins)
                sheet['K' + str(row)] = float(elo_ranking)

            if loser == player:
                sheet['J' + str(row)] = int(amount_of_wins)
                sheet['L' + str(row)] = float(elo_ranking)

            if sheet['J' + str(row)].value is not None and sheet['L' + str(row)].value is not None and \
                    sheet['I' + str(row)].value is not None and sheet['K' + str(row)].value is not None:
                break

    os.chdir('E:/practice/AtpPredictionBot/Generated Data')
    workbook.save('atp_data_f.xlsx')

    workbook.close()
    features_file.close()


def remove_rows_with_blank_values():
    """
    Delete row if it has at least one blank cell
    """
    atp_data_path = 'E:/practice/AtpPredictionBot/Generated Data'
    excel_file = Path(atp_data_path, 'atp_data.xlsx')
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    rows = sheet.max_row

    out_workbook = openpyxl.Workbook()
    out_sheet = out_workbook.active

    for row in range(1, rows+1):
        winner_coef = sheet['E' + str(row)].value
        loser_coef = sheet['F' + str(row)].value
        if winner_coef is None or loser_coef is None:
            continue
        else:
            id = sheet['A' + str(row)].value
            date = sheet['B' + str(row)].value
            winner = sheet['C' + str(row)].value
            loser = sheet['D' + str(row)].value
            w_coef = winner_coef
            l_coef = loser_coef
            w_rank = sheet['G' + str(row)].value
            l_rank = sheet['H' + str(row)].value
            values = [id, date, winner, loser, w_coef, l_coef, w_rank, l_rank]
            out_sheet.append(values)

    os.chdir('E:/practice/AtpPredictionBot/Generated Data')
    out_workbook.save('atp_data_f.xlsx')
    out_workbook.close()
    workbook.close()


def recalculate_total_amount_of_wins():
    """
    Recalculate total amount of wins for each player,
    starting from first match -> players have 0 amount of wins,
    then +/-1 depending if previous match they won/lost.
    """
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    atp_data_path = 'E:/practice/AtpPredictionBot/Generated Data'
    excel_file = Path(atp_data_path, 'atp_data.xlsx')
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    rows = sheet.max_row

    sheet['G' + str(2)].value = 0
    sheet['H' + str(2)].value = 0

    out_id = sheet['A' + str(2)].value
    out_date = sheet['B' + str(2)].value
    out_winner = sheet['C' + str(2)].value
    out_loser = sheet['D' + str(2)].value
    out_wrank = sheet['E' + str(2)].value
    out_lrank = sheet['F' + str(2)].value
    out_wwins = sheet['G' + str(2)].value
    out_lwins = sheet['H' + str(2)].value
    out_welo = sheet['I' + str(2)].value
    out_lelo = sheet['J' + str(2)].value

    row_values = [out_id, out_date, out_winner, out_loser, out_wrank, out_lrank, out_wwins, out_lwins, out_welo,
                  out_lelo]
    output_sheet.append(row_values)

    for row in range(2, rows+1):
        winner = sheet['C' + str(row)].value
        loser = sheet['D' + str(row)].value

        winner_wins_set = False
        loser_wins_set = False

        for previous_row in range(row-1, 1, -1):
            prev_winner = sheet['C' + str(previous_row)].value
            prev_loser = sheet['D' + str(previous_row)].value

            # set player's amount of wins as previous +/- 1
            if winner == prev_winner:
                prev_w_wins = sheet['G' + str(previous_row)].value
                sheet['G' + str(row)].value = prev_w_wins + 1
                winner_wins_set = True
            elif winner == prev_loser:
                prev_l_wins = sheet['H' + str(previous_row)].value
                sheet['G' + str(row)].value = prev_l_wins - 1
                winner_wins_set = True

            if loser == prev_loser:
                prev_l_wins = sheet['H' + str(previous_row)].value
                sheet['H' + str(row)].value = prev_l_wins - 1
                loser_wins_set = True
            elif loser == prev_winner:
                prev_w_wins = sheet['G' + str(previous_row)].value
                sheet['H' + str(row)].value = prev_w_wins + 1
                loser_wins_set = True

            # exit from loop if amount of wins changed or rows finished
            if loser_wins_set and winner_wins_set:
                out_id = sheet['A' + str(row)].value
                out_date = sheet['B' + str(row)].value
                out_winner = sheet['C' + str(row)].value
                out_loser = sheet['D' + str(row)].value
                out_wrank = sheet['E' + str(row)].value
                out_lrank = sheet['F' + str(row)].value
                out_wwins = sheet['G' + str(row)].value
                out_lwins = sheet['H' + str(row)].value
                out_welo = sheet['I' + str(row)].value
                out_lelo = sheet['J' + str(row)].value

                row_values = [out_id, out_date, out_winner, out_loser, out_wrank, out_lrank, out_wwins, out_lwins, out_welo, out_lelo]
                output_sheet.append(row_values)
                break

            # if it is last iteration then set default amount of wins as zero for both players
            if previous_row == 2:
                sheet['G' + str(row)].value = 0
                sheet['H' + str(row)].value = 0

                out_id = sheet['A' + str(row)].value
                out_date = sheet['B' + str(row)].value
                out_winner = sheet['C' + str(row)].value
                out_loser = sheet['D' + str(row)].value
                out_wrank = sheet['E' + str(row)].value
                out_lrank = sheet['F' + str(row)].value
                out_wwins = sheet['G' + str(row)].value
                out_lwins = sheet['H' + str(row)].value
                out_welo = sheet['I' + str(row)].value
                out_lelo = sheet['J' + str(row)].value

                row_values = [out_id, out_date, out_winner, out_loser, out_wrank, out_lrank, out_wwins, out_lwins, out_welo, out_lelo]
                output_sheet.append(row_values)
                break

    os.chdir('E:/practice/AtpPredictionBot/Generated Data')
    output_workbook.save('new_atp_data.xlsx')
    output_workbook.close()
    workbook.close()


def split_data_2_lines_per_match():
    """
    Format train dataset into 2 lines per match:
    first line - match data as it is, and Y would be 1 (winner)
    second line - switch winner and loser parameters and define Y as 0 (loser)
    """

    # input raw data
    atp_data_path = 'E:/practice/AtpPredictionBot/Generated Data'
    excel_file = Path(atp_data_path, 'atp_data.xlsx')
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    rows = sheet.max_row

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    row_id = 1

    for row in range(2, rows + 1):
        try:
            w_coef = float(sheet['E' + str(row)].value)
            l_coef = float(sheet['F' + str(row)].value)
            w_rank = int(sheet['G' + str(row)].value)
            l_rank = int(sheet['H' + str(row)].value)
            w_wins = int(sheet['I' + str(row)].value)
            l_wins = int(sheet['J' + str(row)].value)
            w_elo = float(sheet['K' + str(row)].value)
            l_elo = float(sheet['L' + str(row)].value)
        except:
            continue

        data = [row_id, w_coef, l_coef, w_rank, l_rank, w_wins, l_wins, w_elo, l_elo, 1]
        row_id += 1
        output_sheet.append(data)

        reverse_data = [row_id, l_coef, w_coef, l_rank, w_rank, l_wins, w_wins, l_elo, w_elo, 0]
        row_id += 1
        output_sheet.append(reverse_data)

    os.chdir('E:/practice/AtpPredictionBot/Generated Data')
    output_workbook.save('atp_data_train.xlsx')
    output_workbook.close()
    workbook.close()


print('Program has been started...')
split_data_2_lines_per_match()
print('Program stopped.')
